from datetime import date as date_type
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models.ogl import OGL
from models.ogl_analysis import OGLAnalysis
from models.user import User
from schemas.level_register import (
    OGLAnalysisListResponse,
    OGLAnalysisPagedResponse,
    OGLAnalysisRow,
    OGLListResponse,
)

router = APIRouter()


@router.get("/", response_model=OGLListResponse, summary="List OGL records")
def list_ogl(
    project_id: str = Query(...),
    chainage_from: Optional[int] = Query(None),
    chainage_to: Optional[int] = Query(None),
    road_side: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(OGL).filter(
        OGL.project_id == project_id,
        OGL.is_active == True,
    )
    if chainage_from is not None:
        q = q.filter(OGL.chainage >= chainage_from)
    if chainage_to is not None:
        q = q.filter(OGL.chainage <= chainage_to)
    if road_side:
        q = q.filter(OGL.road_side == road_side.upper())

    total = q.count()
    entries = q.order_by(OGL.chainage.asc(), OGL.road_side.asc()).offset(skip).limit(limit).all()
    return OGLListResponse(total=total, entries=entries)


@router.get(
    "/download/",
    summary="Download OGL data as formatted Excel",
)
def download_ogl(
    project_id: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records = (
        db.query(OGL)
        .filter(OGL.project_id == project_id, OGL.is_active == True)
        .order_by(OGL.chainage.asc(), OGL.road_side.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OGL"

    HEADERS = [
        "Chainage", "Side", "OGL CL RL", "FRL Center",
        "RL@2m", "RL@6m", "RL@Edge", "Version", "Uploaded At",
    ]
    HDR_FILL = PatternFill("solid", fgColor="1E293B")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")

    ws.append(HEADERS)
    for ci in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for ri, r in enumerate(records, 2):
        ws.append([
            r.chainage,
            r.road_side,
            float(r.ogl_cl) if r.ogl_cl is not None else None,
            float(r.frl_center) if r.frl_center is not None else None,
            float(r.rl_at_2m) if r.rl_at_2m is not None else None,
            float(r.rl_at_6m) if r.rl_at_6m is not None else None,
            float(r.rl_at_edge) if r.rl_at_edge is not None else None,
            r.version,
            r.uploaded_at.strftime("%Y-%m-%d %H:%M") if r.uploaded_at else None,
        ])
        if ri % 2 == 0:
            for ci in range(1, len(HEADERS) + 1):
                ws.cell(row=ri, column=ci).fill = ALT_FILL

    for ci, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{project_id}_OGL_{date_type.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/analysis/at-chainage/",
    response_model=List[OGLAnalysisRow],
    summary="Get cut/fill analysis for both sides at a specific chainage",
)
def get_analysis_at_chainage(
    project_id: str = Query(...),
    chainage: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rows = (
        db.query(OGLAnalysis)
        .filter(
            OGLAnalysis.project_id == project_id,
            OGLAnalysis.chainage == chainage,
            OGLAnalysis.is_active == True,
        )
        .order_by(OGLAnalysis.road_side.asc())
        .all()
    )
    return rows


@router.get(
    "/analysis/paged/",
    response_model=OGLAnalysisPagedResponse,
    summary="List OGL cut/fill analysis with pagination",
)
def list_analysis_paged(
    project_id: str = Query(...),
    road_side: Optional[str] = Query(None),
    cut_fill_type: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    base_q = db.query(OGLAnalysis).filter(
        OGLAnalysis.project_id == project_id,
        OGLAnalysis.is_active == True,
    )
    if road_side:
        base_q = base_q.filter(OGLAnalysis.road_side == road_side.upper())

    cut = base_q.filter(OGLAnalysis.cut_fill_type == "CUT").count()
    fill = base_q.filter(OGLAnalysis.cut_fill_type == "FILL").count()
    zero = base_q.filter(OGLAnalysis.cut_fill_type == "ZERO").count()

    paged_q = base_q
    if cut_fill_type:
        paged_q = paged_q.filter(OGLAnalysis.cut_fill_type == cut_fill_type.upper())

    total = paged_q.count()
    entries = (
        paged_q.order_by(OGLAnalysis.chainage.asc(), OGLAnalysis.road_side.asc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return OGLAnalysisPagedResponse(
        total=total,
        cut_chainages=cut,
        fill_chainages=fill,
        zero_chainages=zero,
        entries=entries,
    )


@router.get(
    "/analysis/",
    response_model=OGLAnalysisListResponse,
    summary="List OGL cut/fill analysis with summary counts",
)
def list_analysis(
    project_id: str = Query(...),
    road_side: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(OGLAnalysis).filter(
        OGLAnalysis.project_id == project_id,
        OGLAnalysis.is_active == True,
    )
    if road_side:
        q = q.filter(OGLAnalysis.road_side == road_side.upper())

    entries = q.order_by(OGLAnalysis.chainage.asc(), OGLAnalysis.road_side.asc()).all()

    cut = sum(1 for e in entries if e.cut_fill_type == "CUT")
    fill = sum(1 for e in entries if e.cut_fill_type == "FILL")
    zero = sum(1 for e in entries if e.cut_fill_type == "ZERO")

    return OGLAnalysisListResponse(
        total=len(entries),
        entries=entries,
        cut_chainages=cut,
        fill_chainages=fill,
        zero_chainages=zero,
    )


@router.get(
    "/analysis/download/",
    summary="Download OGL Analysis as 2-sheet formatted Excel",
)
def download_analysis(
    project_id: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records = (
        db.query(OGLAnalysis)
        .filter(OGLAnalysis.project_id == project_id, OGLAnalysis.is_active == True)
        .order_by(OGLAnalysis.chainage.asc(), OGLAnalysis.road_side.asc())
        .all()
    )

    cut_count  = sum(1 for r in records if r.cut_fill_type == "CUT")
    fill_count = sum(1 for r in records if r.cut_fill_type == "FILL")
    zero_count = sum(1 for r in records if r.cut_fill_type == "ZERO")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ──────────────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "OGL Analysis"

    DATA_HEADERS = [
        "Chainage", "Side", "OGL RL", "EMB FRL",
        "Cut/Fill (m)", "Type", "Cross Area (m²)", "Volume (m³)", "Computed At",
    ]
    HDR_FILL  = PatternFill("solid", fgColor="1E293B")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    CUT_FILL  = PatternFill("solid", fgColor="FEE2E2")
    FILL_FILL = PatternFill("solid", fgColor="DCFCE7")
    ZERO_FILL = PatternFill("solid", fgColor="F1F5F9")

    ws_data.append(DATA_HEADERS)
    for ci in range(1, len(DATA_HEADERS) + 1):
        cell = ws_data.cell(row=1, column=ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_data.freeze_panes = "A2"

    for ri, r in enumerate(records, 2):
        ws_data.append([
            r.chainage,
            r.road_side,
            float(r.ogl_rl) if r.ogl_rl is not None else None,
            float(r.emb_frl) if r.emb_frl is not None else None,
            float(r.cut_fill_m) if r.cut_fill_m is not None else None,
            r.cut_fill_type,
            float(r.cross_area_sqm) if r.cross_area_sqm is not None else None,
            float(r.volume_cum) if r.volume_cum is not None else None,
            r.computed_at.strftime("%Y-%m-%d %H:%M") if r.computed_at else None,
        ])
        row_fill = (
            CUT_FILL  if r.cut_fill_type == "CUT"  else
            FILL_FILL if r.cut_fill_type == "FILL" else
            ZERO_FILL
        )
        for ci in range(1, len(DATA_HEADERS) + 1):
            ws_data.cell(row=ri, column=ci).fill = row_fill

    for ci, col_cells in enumerate(ws_data.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws_data.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "1E3A5F"

    chainages = sorted({r.chainage for r in records})
    ch_min = chainages[0] if chainages else None
    ch_max = chainages[-1] if chainages else None

    summary_rows = [
        ["Project ID",           project_id],
        ["Generated At",         date_type.today().isoformat()],
        ["Total Chainages",      len(chainages)],
        ["Chainage Range",       f"{ch_min} – {ch_max}" if ch_min else "—"],
        ["Total Records",        len(records)],
        ["CUT Chainages",        cut_count],
        ["FILL Chainages",       fill_count],
        ["ZERO Chainages",       zero_count],
    ]
    KEY_FONT  = Font(bold=True, size=10)
    for row_data in summary_rows:
        ws_sum.append(row_data)
    for row in ws_sum.iter_rows(min_row=1, max_row=len(summary_rows)):
        row[0].font = KEY_FONT
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{project_id}_OGL_Analysis_{date_type.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
