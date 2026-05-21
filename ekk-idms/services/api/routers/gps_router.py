from datetime import date as date_type
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models.gps_coordinates import GPSCoordinates
from models.user import User
from schemas.level_register import GPSListResponse, GPSRow

router = APIRouter()


@router.get("/", response_model=GPSListResponse, summary="List GPS coordinate records")
def list_gps(
    project_id: str = Query(...),
    chainage_from: Optional[int] = Query(None),
    chainage_to: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(GPSCoordinates).filter(GPSCoordinates.project_id == project_id)
    if chainage_from is not None:
        q = q.filter(GPSCoordinates.chainage_from >= chainage_from)
    if chainage_to is not None:
        q = q.filter(GPSCoordinates.chainage_to <= chainage_to)

    entries = q.order_by(GPSCoordinates.chainage_from.asc()).all()
    return GPSListResponse(total=len(entries), entries=entries)


@router.get(
    "/download/",
    summary="Download GPS coordinates as formatted Excel",
)
def download_gps(
    project_id: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records = (
        db.query(GPSCoordinates)
        .filter(GPSCoordinates.project_id == project_id)
        .order_by(GPSCoordinates.chainage_from.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GPS Coordinates"
    ws.sheet_properties.tabColor = "0EA5E9"

    HEADERS = [
        "CH From", "CH To", "NH Number", "State", "District", "RO", "PIU",
        "Lat Start", "Lon Start", "Alt Start (m)",
        "Lat End", "Lon End", "Alt End (m)",
        "Uploaded At",
    ]
    HDR_FILL = PatternFill("solid", fgColor="1E293B")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")

    ws.append(HEADERS)
    for ci in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for ri, r in enumerate(records, 2):
        ws.append([
            r.chainage_from,
            r.chainage_to,
            r.nh_number,
            r.state,
            r.district,
            r.ro,
            r.piu,
            float(r.lat_start) if r.lat_start is not None else None,
            float(r.lon_start) if r.lon_start is not None else None,
            float(r.alt_start_m) if r.alt_start_m is not None else None,
            float(r.lat_end) if r.lat_end is not None else None,
            float(r.lon_end) if r.lon_end is not None else None,
            float(r.alt_end_m) if r.alt_end_m is not None else None,
            r.uploaded_at.strftime("%Y-%m-%d %H:%M") if r.uploaded_at else None,
        ])
        if ri % 2 == 0:
            for ci in range(1, len(HEADERS) + 1):
                ws.cell(row=ri, column=ci).fill = ALT_FILL

    for ci, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 20)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{project_id}_GPS_{date_type.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/at-chainage/",
    response_model=GPSRow,
    summary="Find GPS record containing a specific chainage",
)
def get_at_chainage(
    project_id: str = Query(...),
    chainage: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    record = (
        db.query(GPSCoordinates)
        .filter(
            GPSCoordinates.project_id == project_id,
            GPSCoordinates.chainage_from <= chainage,
            GPSCoordinates.chainage_to >= chainage,
        )
        .first()
    )
    if not record:
        raise HTTPException(
            status_code=404,
            detail=f"No GPS record covers chainage {chainage} for project {project_id}",
        )
    return record
