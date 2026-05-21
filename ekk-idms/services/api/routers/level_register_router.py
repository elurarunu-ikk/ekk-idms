import uuid
from collections import defaultdict
from datetime import date as date_type
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models.level_register import LevelRegister
from models.user import User
from schemas.level_register import (
    LevelRegisterListResponse,
    LevelRegisterRow,
    LevelRegisterSummaryResponse,
    LevelRegisterSummaryRow,
)

router = APIRouter()

LAYER_ORDER = ["EMB", "SG", "GSB", "CTB", "WMM", "DBM", "BC"]
_TAB_COLORS = {
    "EMB": "00B050",
    "SG":  "FFC000",
    "GSB": "7030A0",
    "CTB": "FF0000",
    "WMM": "00B0F0",
    "DBM": "FF6600",
    "BC":  "0070C0",
}


@router.get(
    "/summary/",
    response_model=LevelRegisterSummaryResponse,
    summary="Get per-layer summary for a project",
)
def get_summary(
    project_id: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rows = (
        db.query(
            LevelRegister.layer_code,
            func.count(LevelRegister.id).label("total_records"),
            func.min(LevelRegister.chainage).label("chainage_min"),
            func.max(LevelRegister.chainage).label("chainage_max"),
            func.min(LevelRegister.frl_center).label("frl_min"),
            func.max(LevelRegister.frl_center).label("frl_max"),
        )
        .filter(
            LevelRegister.project_id == project_id,
            LevelRegister.is_active == True,
        )
        .group_by(LevelRegister.layer_code)
        .all()
    )

    layers = [
        LevelRegisterSummaryRow(
            layer_code=r.layer_code,
            total_records=r.total_records,
            chainage_min=r.chainage_min,
            chainage_max=r.chainage_max,
            frl_min=float(r.frl_min) if r.frl_min is not None else None,
            frl_max=float(r.frl_max) if r.frl_max is not None else None,
        )
        for r in rows
    ]
    return LevelRegisterSummaryResponse(project_id=project_id, layers=layers)


@router.get(
    "/download/",
    summary="Download Level Register as multi-sheet formatted Excel",
)
def download_level_register(
    project_id: str = Query(...),
    layer_code: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    q = db.query(LevelRegister).filter(
        LevelRegister.project_id == project_id,
        LevelRegister.is_active == True,
    )
    if layer_code:
        q = q.filter(LevelRegister.layer_code == layer_code.upper())
    records = q.order_by(
        LevelRegister.layer_code.asc(),
        LevelRegister.chainage.asc(),
        LevelRegister.road_side.asc(),
    ).all()

    by_layer: dict = defaultdict(list)
    for r in records:
        by_layer[r.layer_code].append(r)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADERS = [
        "Chainage", "Side", "Layer", "Description", "Thickness(mm)",
        "FRL Center", "Camber%", "Camber Type", "Road Width(m)",
        "RL@0m", "RL@2m", "RL@6m", "RL@9.5m", "RL@11m", "RL@Edge",
        "TCS Ref", "Version", "Uploaded At",
    ]
    HDR_FILL = PatternFill("solid", fgColor="1E293B")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    CENTER = Alignment(horizontal="center", vertical="center")

    layers_to_render = [lc for lc in LAYER_ORDER if lc in by_layer]
    for lc in by_layer:
        if lc not in layers_to_render:
            layers_to_render.append(lc)

    if not layers_to_render:
        ws = wb.create_sheet("No Data")
        ws.append(["No data found for the given filters."])
    else:
        for lc in layers_to_render:
            ws = wb.create_sheet(lc)
            ws.sheet_properties.tabColor = _TAB_COLORS.get(lc, "808080")
            ws.append(HEADERS)
            for ci in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 20

            for ri, r in enumerate(by_layer[lc], 2):
                row_data = [
                    r.chainage,
                    r.road_side,
                    r.layer_code,
                    r.layer_desc,
                    r.thickness_mm,
                    float(r.frl_center) if r.frl_center is not None else None,
                    float(r.camber_pct) if r.camber_pct is not None else None,
                    r.camber_type,
                    float(r.road_width_m) if r.road_width_m is not None else None,
                    float(r.rl_at_0m) if r.rl_at_0m is not None else None,
                    float(r.rl_at_2m) if r.rl_at_2m is not None else None,
                    float(r.rl_at_6m) if r.rl_at_6m is not None else None,
                    float(r.rl_at_9_5m) if r.rl_at_9_5m is not None else None,
                    float(r.rl_at_11m) if r.rl_at_11m is not None else None,
                    float(r.rl_at_edge) if r.rl_at_edge is not None else None,
                    r.tcs_ref,
                    r.version,
                    r.uploaded_at.strftime("%Y-%m-%d %H:%M") if r.uploaded_at else None,
                ]
                ws.append(row_data)
                if ri % 2 == 0:
                    for ci in range(1, len(HEADERS) + 1):
                        ws.cell(row=ri, column=ci).fill = ALT_FILL

            for ci, col_cells in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                )
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    suffix = f"_{layer_code.upper()}" if layer_code else ""
    filename = f"{project_id}_Level_Register{suffix}_{date_type.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/at-chainage/",
    response_model=LevelRegisterRow,
    summary="Get level data at a specific chainage for validation",
)
def get_at_chainage(
    project_id: str = Query(...),
    chainage: int = Query(...),
    layer_code: str = Query(...),
    road_side: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    record = (
        db.query(LevelRegister)
        .filter(
            LevelRegister.project_id == project_id,
            LevelRegister.chainage == chainage,
            LevelRegister.layer_code == layer_code.upper(),
            LevelRegister.road_side == road_side.upper(),
            LevelRegister.is_active == True,
        )
        .first()
    )
    if not record:
        raise HTTPException(
            status_code=404,
            detail=f"No level register entry for project={project_id} layer={layer_code} CH={chainage} side={road_side}",
        )
    return record


@router.get(
    "/{record_id}",
    response_model=LevelRegisterRow,
    summary="Get a single level register record by ID",
)
def get_by_id(
    record_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    record = db.query(LevelRegister).filter(LevelRegister.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Level register record not found")
    return record


@router.get(
    "/",
    response_model=LevelRegisterListResponse,
    summary="List level register entries with optional filters",
)
def list_entries(
    project_id: str = Query(...),
    layer_code: Optional[str] = Query(None),
    chainage_from: Optional[int] = Query(None),
    chainage_to: Optional[int] = Query(None),
    road_side: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, le=1000),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(LevelRegister).filter(
        LevelRegister.project_id == project_id,
        LevelRegister.is_active == True,
    )
    if layer_code:
        q = q.filter(LevelRegister.layer_code == layer_code.upper())
    if chainage_from is not None:
        q = q.filter(LevelRegister.chainage >= chainage_from)
    if chainage_to is not None:
        q = q.filter(LevelRegister.chainage <= chainage_to)
    if road_side:
        q = q.filter(LevelRegister.road_side == road_side.upper())

    total = q.count()
    entries = q.order_by(LevelRegister.chainage.asc(), LevelRegister.layer_code.asc()).offset(skip).limit(limit).all()
    return LevelRegisterListResponse(total=total, entries=entries)
